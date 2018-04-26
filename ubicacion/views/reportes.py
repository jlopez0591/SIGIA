from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import PatternFill, colors, Font
from ubicacion.models import FacultadInstancia, EscuelaInstancia, DepartamentoInstancia
import datetime


def reporte_facultad(request, facultad_pk):
    facultad = FacultadInstancia.objects.get(pk=facultad_pk)
    # Generacion de Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = 'Registro de Unidad'

    # Declaracion de Paginas Adicionales
    ws1 = wb.create_sheet('Inventario')
    ws2 = wb.create_sheet('Profesores')
    ws3 = wb.create_sheet('Escuelas')

    # # # Declaracion de Formatos
    title_font = Font(size=15, bold=True, color=colors.WHITE)
    subtitle_font = Font(size=13, bold=True, color=colors.WHITE)
    datacell_font = Font(italic=True, size=11)

    title_fill = PatternFill(start_color='FF4295f4', end_color='FF4295f4', fill_type='solid')

    # Asignacion de Formatos
    ws['A1'].font = title_font
    ws['A2'].font = subtitle_font
    ws['A5'].font = datacell_font
    ws['A7'].font = datacell_font
    ws['A8'].font = datacell_font
    ws['A9'].font = datacell_font
    ws['A10'].font = datacell_font
    ws['A11'].font = datacell_font

    ws['A1'].fill = title_fill
    ws['A2'].fill = title_fill

    # Datos de las Celda
    # Reporte General

    # Titulos
    # ws.merge_cells('A1:B1')
    # ws.merge_cells('A2:B2')
    # ws.merge_cells('A5:B5')
    ws['A1'] = 'Universidad de Panama'
    ws.column_dimensions['A'].width = 50.0
    ws['A2'] = '{0}'.format(facultad.facultad.nombre)
    ws['A5'] = 'Informe General - 2018'
    ws['A7'] = 'Cantidad de Profesores Activos'
    ws['A8'] = 'Cantidad de Estudiantes Registrados'
    ws['A9'] = 'Cantidad de Recursos Fisicos'
    ws['A10'] = 'Cantidad de Anteproyectos Entregados'
    ws['A11'] = 'Cantidad de Trabajos de Graduacion Sustentados'

    # Datos
    ws['B7'] = facultad.personal.profesores().count()
    ws['B8'] = facultad.estudiantes.count()
    ws['B9'] = facultad.equipos.all().count()
    ws['B10'] = facultad.trabajos_graduacion.all().count()
    ws['B11'] = facultad.trabajos_graduacion.filter(estado='sustentado').count()

    # Inventario

    ## Encabezado
    ws1['A1'] = 'Universidad de Panama'
    ws1.column_dimensions['A'].width = 50.0
    ws1['A2'] = 'Facultad de Prueba'
    ws1['A5'] = 'Informe de Inventario - 2018'

    ws1['A1'].font = title_font
    ws1['A2'].font = subtitle_font
    ws1['A5'].font = datacell_font

    ws1['A1'].fill = title_fill
    ws1['A2'].fill = title_fill

    ## Titulos
    ws1['A7'] = 'Costo Total de Recuros Fisicos'

    ### TODO: Iterar y generar celdas por tipo.

    # Datos
    ws1['B7'] = 0

    # Profesores
    ws2['A1'] = 'Universidad de Panama'
    ws2.column_dimensions['A'].width = 50.0
    ws2['A2'] = 'Facultad de Prueba'
    ws2['A5'] = 'Informe de Profesores - 2018'
    ws2['A1'].font = title_font
    ws2['A2'].font = subtitle_font
    ws2['A5'].font = datacell_font

    ws2['A1'].fill = title_fill
    ws2['A2'].fill = title_fill
    ws2['A7'] = 'Cantidad de Profesores con Doctorado'
    ws2['A8'] = 'Cantidad de Profesores con Maestria'
    ws2['A9'] = 'Cantidad de Profesores con Postgrado'
    ws2['A10'] = 'Cantidad de Profesores con Licenciatura'

    profesores_titulos = {
        'doctorado': 0,
        'maestria': 0,
        'postgrado': 0,
        'licenciado': 0
    }
    for profesor in facultad.personal.profesores():
        nivel_profesor = profesor.nivel_academico()
        profesores_titulos[nivel_profesor.lower()] += 1

    ws2['B7'] = profesores_titulos['doctorado']
    ws2['B8'] = profesores_titulos['maestria']
    ws2['B9'] = profesores_titulos['postgrado']
    ws2['B10'] = profesores_titulos['licenciado']

    ws2['A1'].font = title_font
    ws2['A2'].font = subtitle_font
    ws2['A5'].font = datacell_font

    # Ciclo para cantidad por departamento
    # Ciclo por categoria
    # Ciclo por dedicacion

    # Estudiantes
    ws3['A1'] = 'Universidad de Panama'
    ws3.column_dimensions['A'].width = 50.0
    ws3['A2'] = 'Facultad de Prueba'
    ws3['A5'] = 'Informe de Escuelas - 2018'

    ws3['A1'].font = title_font
    ws3['A2'].font = subtitle_font
    ws3['A5'].font = datacell_font
    ws3['A1'].fill = title_fill
    ws3['A2'].fill = title_fill
    # Ciclo de Estudiantes por Escuela

    info_escuelas = dict()

    for escuela in facultad.escuelas.all():
        info_escuelas[escuela.escuela.nombre] = {}
        info_escuelas[escuela.escuela.nombre]['estudiantes'] = escuela.estudiantes.all().count()
        info_escuelas[escuela.escuela.nombre]['anteproyectos'] = escuela.estudiantes.all().count()
        info_escuelas[escuela.escuela.nombre]['trabajos'] = escuela.estudiantes.all().count()

    ca = 7
    for key, value in info_escuelas.items():
        ws3['A' + str(ca)] = 'Cantidad de Estudiantes en {0}'.format(key)
        ws3['B' + str(ca)] = value['estudiantes']
        ca += 1
        ws3['A' + str(ca)] = 'Cantidad de Anteproyectos en {0}'.format(key)
        ws3['B' + str(ca)] = value['anteproyectos']
        ca += 1
        ws3['A' + str(ca)] = 'Cantidad de Trabajos de Graduacion en {0}'.format(key)
        ws3['B' + str(ca)] = value['trabajos']
        ca += 1

    # Ciclo Anteproyectos por Escuela
    # Ciclo Trabajos de Graduacion por Escuela

    fecha_actual = datetime.datetime.now().strftime("%Y-%m")
    nombre_facultad = facultad.facultad.nombre
    filename = '{0}-{1}'.format(fecha_actual, nombre_facultad)
    response = HttpResponse(content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename=reporte_{0}.xlsx'.format(filename)

    wb.save(response)

    return response


def reporte_escuela(request, escuela_pk):
    escuela = EscuelaInstancia.objects.get(pk=escuela_pk)
    # Vistazo General - Cantidad de Estudiantes, Anteproyectos Entregados, Trabajos Sustentados.
    # Numero de Estudiantes por carrera
    carreras = escuela.carreras.all()
    estudiantes_carrera = dict()
    for carrera in carreras:
        estudiantes_carrera[carrera] = carrera.estudiantes.all().count()

    # Anteproyectos entregados por Nivel
    anteproyectos = escuela.trabajos_graduacion.aprobados()
    dict_anteproyectos = dict()
    for anteproyecto in anteproyectos:
        programa = anteproyecto.get_programa_display()
        if programa in dict_anteproyectos:
            dict_anteproyectos[programa] += 1
        else:
            dict_anteproyectos[programa] = 0

    # Trabajos Sustentados por Nivel
    trabajos = escuela.trabajos_graduacion.sustentados()
    dict_trabajos = dict()
    for trabajo in trabajos:
        programa = trabajo.get_programa_display()
        if programa in dict_trabajos:
            dict_trabajos[programa] += 1
        else:
            dict_trabajos[programa] = 0

    fecha_actual = datetime.datetime.now().strftime("%Y-%m")
    nombre_escuela = escuela.escuela.nombre

    wb = Workbook()
    ws = wb.active
    ws.title = 'Informe de Escuela'

    title_font = Font(size=15, bold=True, color=colors.WHITE)
    subtitle_font = Font(size=13, bold=True, color=colors.WHITE)
    datacell_font = Font(italic=True, size=11)

    title_fill = PatternFill(start_color='FF4295f4', end_color='FF4295f4', fill_type='solid')

    ws['A1'].font = title_font
    ws['A2'].font = subtitle_font
    ws['A5'].font = datacell_font
    ws['A7'].font = datacell_font
    ws['A8'].font = datacell_font
    ws['A9'].font = datacell_font
    ws['A10'].font = datacell_font
    ws['A11'].font = datacell_font

    ws['A1'].fill = title_fill
    ws['A2'].fill = title_fill

    ws['A1'] = 'Universidad de Panama'
    ws.column_dimensions['A'].width = 50.0
    ws['A2'] = escuela.escuela.nombre
    ws['A5'] = 'Informe General'

    ws['A7'] = 'Estudiantes Activos'
    ws['B7'] = escuela.estudiantes.activos().count()
    ws['A8'] = 'Anteproyectos Registrados'
    ws['B8'] = escuela.trabajos_graduacion.all().count()
    ws['A9'] = 'Trabajos de Graduacion Sustentados'
    ws['B9'] = escuela.trabajos_graduacion.sustentados().count()

    i = 12  # Indice de la celda
    for carrera in escuela.carreras.all():
        ws['A' + str(i)] = 'Estudiantes en {0}'.format(carrera)
        ws['B' + str(i)] = carrera.estudiantes.activos().count()
        i += 1
        ws['A' + str(i)] = 'Anteproyectos Entregados en {0}'.format(carrera)
        ws['B' + str(i)] = carrera.trabajos_graduacion.aprobados().count()
        i += 1
        ws['A' + str(i)] = 'Trabajos Sustentados en {0}'.format(carrera)
        ws['B' + str(i)] = carrera.trabajos_graduacion.sustentados().count()
        i += 1

    filename = '{0}-{1}'.format(fecha_actual, nombre_escuela)
    response = HttpResponse(content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename=reporte_{0}.xlsx'.format(filename)
    wb.save(response)
    return response


def reporte_departamento(request, departamento_pk):
    departamento = DepartamentoInstancia.objects.get(pk=departamento_pk)
    wb = Workbook()
    ws = wb.active
    ws.title = 'Informe de Departamento'

    # Vistazo General - Cantidad de Profesores por Dedicacion, Categoria, Nivel Academico, Actividades por Tipo y aprobadas.

    fecha_actual = datetime.datetime.now().strftime("%Y-%m")
    nombre_departamento = departamento.departamento.nombre
    filename = '{0}-{1}'.format(fecha_actual, nombre_departamento)
    response = HttpResponse(content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename=reporte_{0}.xlsx'.format(filename)
    wb.save(response)
    return response


def reporte_facultad_demo(request):
    # Generacion de Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = 'Registro de Unidad'

    # Declaracion de Paginas Adicionales
    ws1 = wb.create_sheet('Inventario')
    ws2 = wb.create_sheet('Profesores')
    ws3 = wb.create_sheet('Escuelas')

    # # # Declaracion de Formatos
    title_font = Font(size=15, bold=True, color=colors.WHITE)
    subtitle_font = Font(size=13, bold=True, color=colors.WHITE)
    datacell_font = Font(italic=True, size=11)

    title_fill = PatternFill(start_color='FF4295f4', end_color='FF4295f4', fill_type='solid')

    # Asignacion de Formatos
    ws['A1'].font = title_font
    ws['A2'].font = subtitle_font
    ws['A5'].font = datacell_font
    ws['A7'].font = datacell_font
    ws['A8'].font = datacell_font
    ws['A9'].font = datacell_font
    ws['A10'].font = datacell_font
    ws['A11'].font = datacell_font

    ws['A1'].fill = title_fill
    ws['A2'].fill = title_fill

    # Datos de las Celda
    # Reporte General

    # Titulos
    ws['A1'] = 'Universidad de Panama'
    ws['A2'] = 'Facultad de Prueba'
    ws['A5'] = 'Informe General - 2018'
    ws['A7'] = 'Cantidad de Profesores Activos'
    ws['A8'] = 'Cantidad de Estudiantes Registrados'
    ws['A9'] = 'Cantidad de Recursos Fisicos'
    ws['A10'] = 'Cantidad de Anteproyectos Entregados'
    ws['A11'] = 'Cantidad de Trabajos de Graduacion Sustentados'

    # Datos
    ws['B7'] = 0
    ws['B8'] = 0
    ws['B9'] = 0
    ws['B10'] = 0
    ws['B11'] = 0

    # Inventario

    ## Encabezado
    ws1['A1'] = 'Universidad de Panama'
    ws1['A2'] = 'Facultad de Prueba'
    ws1['A5'] = 'Informe de Inventario - 2018'

    ws1['A1'].font = title_font
    ws1['A2'].font = subtitle_font
    ws1['A5'].font = datacell_font

    ## Titulos
    ws1['A7'] = 'Cantidad Total de Recursos Fisicos'
    ws1['A8'] = 'Costo Total de Recuros Fisicos'

    ### TODO: Iterar y generar celdas por tipo.

    # Datos
    ws1['B7'] = 0
    ws1['B8'] = 0

    # Profesores
    ws2['A1'] = 'Universidad de Panama'
    ws2['A2'] = 'Facultad de Prueba'
    ws2['A5'] = 'Informe de Profesores - 2018'
    ws2['A7'] = 'Cantidad de Profesores con Doctorado'
    ws2['A8'] = 'Cantidad de Profesores con Maestria'
    ws2['A9'] = 'Cantidad de Profesores con Postgrado'
    ws2['A10'] = 'Cantidad de Profesores con Licenciatura'

    ws2['A1'].font = title_font
    ws2['A2'].font = subtitle_font
    ws2['A5'].font = datacell_font

    # Ciclo para cantidad por departamento
    # Ciclo por categoria
    # Ciclo por dedicacion

    # Estudiantes
    ws3['A1'] = 'Universidad de Panama'
    ws3['A2'] = 'Facultad de Prueba'
    ws3['A5'] = 'Informe de Escuelas - 2018'

    ws3['A1'].font = title_font
    ws3['A2'].font = subtitle_font
    ws3['A5'].font = datacell_font
    # Ciclo de Estudiantes por Escuela
    celda_actual = 6
    # Ciclo Anteproyectos por Escuela
    # Ciclo Trabajos de Graduacion por Escuela

    filename = 'reporte-prueba'
    response = HttpResponse(content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename={0}.xlsx'.format(filename)

    wb.save(response)

    return response
